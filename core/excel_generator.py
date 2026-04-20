import pandas as pd
import os
from openpyxl.styles import Font, PatternFill, Alignment

def apply_centro_header(worksheet, report_title, num_cols):
    from openpyxl.styles import Font, Alignment
    
    worksheet['A1'] = "CENTRO SERVICES COOPERATIVE"
    worksheet['A2'] = " Purok Camachille, Brgy. Tambler, General Santos City"
    worksheet['A3'] = " centrocooperative21@gmail.com | (083) 554 5552"
    worksheet['A5'] = str(report_title).upper()

    worksheet['A1'].font = Font(bold=True, size=30)
    worksheet['A2'].font = Font(size=11)
    worksheet['A3'].font = Font(size=11)
    worksheet['A5'].font = Font(bold=True, size=25)
    
    merge_end_col = chr(64 + num_cols) if num_cols <= 26 else 'Z'
    if num_cols > 1:
        worksheet.merge_cells(f'A1:{merge_end_col}1')
        worksheet.merge_cells(f'A2:{merge_end_col}2')
        worksheet.merge_cells(f'A3:{merge_end_col}3')
        worksheet.merge_cells(f'A5:{merge_end_col}5')
        
    for i in [1, 2, 3, 5]:
        worksheet[f'A{i}'].alignment = Alignment(horizontal="center")

def generate_excel_report(report_title, table_headers, table_data, filename="report.xlsx"):
    """
    Generates a stylized Excel report from table headers and data.
    """
    df = pd.DataFrame(table_data, columns=table_headers)
    
    # Use ExcelWriter for styling
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Report', startrow=6)
    
    workbook = writer.book
    worksheet = writer.sheets['Report']
    
    apply_centro_header(worksheet, report_title, len(table_headers))
    
    # Basic Styling (Row 7 is now the table header)
    table_header_font = Font(bold=True, color="FFFFFF")
    table_header_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
    
    for cell in worksheet[7]:
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    # Column width adjustment
    from openpyxl.styles import Border, Side
    from openpyxl.utils import get_column_letter
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col_idx, column_cells in enumerate(worksheet.columns, 1):
        # We only care about data rows (7 onwards) for width and borders
        # We skip merged cells gracefully
        data_cells = [c for c in column_cells if hasattr(c, 'row') and c.row >= 7 and hasattr(c, 'value') and c.value is not None]
        
        if data_cells:
            length = max(len(str(cell.value)) for cell in data_cells)
            worksheet.column_dimensions[get_column_letter(col_idx)].width = length + 5
            
        for cell in column_cells:
            # Check if it's a regular cell (not a MergedCell which might lack attributes we need)
            # though MergedCells are usually just placeholders, they can appear in the iterator
            if hasattr(cell, 'row') and cell.row >= 7:
                cell.border = thin_border
        
    writer.close()
    return os.path.abspath(filename)
