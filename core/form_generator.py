import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from core.excel_generator import apply_centro_header

FORM_FILENAME = "BLANK_REQUEST_FORM.xlsx"

def generate_blank_form():
    """Generates a perfectly formatted, printable blank Excel form for supply requests."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Supply Request Form"

    # Define common styles
    bold_font = Font(bold=True)
    header_font = Font(bold=True, size=14, color="FFFFFF")
    title_font = Font(bold=True, size=18)
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Column Widths
    ws.column_dimensions['A'].width = 5   # "#"
    ws.column_dimensions['B'].width = 35  # "Item Name"
    ws.column_dimensions['C'].width = 12  # "Quantity"
    ws.column_dimensions['D'].width = 15  # "Is Refill? (Y/N)"
    ws.column_dimensions['E'].width = 25  # "Frequency / Remarks"

    # Apply standard Centro Header
    apply_centro_header(ws, "MATERIAL / SUPPLY REQUEST FORM", 5)

    # Employee / Department Block (Top) - Shifted down to row 7
    metadata_labels = [
        ("B7", "Date:"), ("D7", "Shift:"),
        ("B8", "Employee Name:"), ("D8", "Employee Role:"),
        ("B9", "Department Area:"), ("D9", "Supervisor:")
    ]
    
    for cell_loc, label in metadata_labels:
        ws[cell_loc] = label
        ws[cell_loc].font = bold_font
        ws[cell_loc].alignment = left_align
        
        # Create a line to write on next to the label
        blank_cell = f"{chr(ord(cell_loc[0]) + 1)}{cell_loc[1]}"
        ws[blank_cell].border = Border(bottom=Side(style='thin'))

    # Table Headers (Row 11) - Shifted from row 8
    headers = ["#", "Item Requested", "Quantity", "Is Refill? (Y/N)", "Frequency / Remarks"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=11, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border

    # Blank Rows for filling (Rows 12 to 28) - Shifted from row 9
    for row_num in range(12, 29):
        # Auto-number the # column
        num_cell = ws.cell(row=row_num, column=1)
        num_cell.value = row_num - 11
        num_cell.alignment = center_align
        num_cell.border = thin_border
        
        # Add borders to blank cells
        for col_num in range(2, 6):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border

    # Signatures block at the bottom
    ws['B28'] = "Requested By:"
    ws['B28'].font = bold_font
    ws['C28'].border = Border(bottom=Side(style='thin'))
    ws.merge_cells('C28:E28')
    
    ws['B30'] = "Approved By:"
    ws['B30'].font = bold_font
    ws['C30'].border = Border(bottom=Side(style='thin'))
    ws.merge_cells('C30:E30')

    # Page Setup for Printing
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    
    # Set print margins
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75

    # Save the form
    wb.save(FORM_FILENAME)
    print(f"Printable form generated at: {os.path.abspath(FORM_FILENAME)}")

def generate_populated_report(employee_name, role, area, shift, supervisor, requests_data):
    """Generates a populated supply request history report for a specific employee."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Request History"

    # Define common styles
    bold_font = Font(bold=True)
    header_font = Font(bold=True, size=14, color="FFFFFF")
    title_font = Font(bold=True, size=18)
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Column Widths
    ws.column_dimensions['A'].width = 15  # "Date"
    ws.column_dimensions['B'].width = 30  # "Item Requested"
    ws.column_dimensions['C'].width = 12  # "Quantity"
    ws.column_dimensions['D'].width = 15  # "Is Refill?"
    ws.column_dimensions['E'].width = 25  # "Frequency"

    # Apply standard Centro Header
    apply_centro_header(ws, f"SUPPLY REQUEST HISTORY: {employee_name}", 5)

    # Metadata Block - Shifted to rows 7-9
    ws['A7'] = "Employee Name:"; ws['B7'] = employee_name
    ws['A8'] = "Employee Role:"; ws['B8'] = role
    ws['A9'] = "Shift:";         ws['B9'] = shift
    ws['D7'] = "Area:";          ws['E7'] = area
    ws['D8'] = "Supervisor:";    ws['E8'] = supervisor or "N/A"

    for row in range(7, 10):
        for col in [1, 4]:
            ws.cell(row=row, column=col).font = bold_font

    # Table Headers (Row 11) - Shifted from row 8
    headers = ["Date", "Item Requested", "Quantity", "Is Refill?", "Frequency / Remarks"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=11, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border

    # Populating Data - starting from Row 12
    for row_idx, data in enumerate(requests_data, 12):
        # data is expected to be (date_str, item_name, qty, refill_str, frequency)
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            cell.alignment = left_align if col_idx == 2 else center_align

    # Page Setup
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    
    filename = f"HISTORY_{employee_name.replace(' ', '_').replace(',', '')}.xlsx"
    wb.save(filename)
    return filename

def generate_consumption_report(data_rows):
    """Generates a summary Excel report for the Consumption & Usage Analysis."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Consumption Analysis"

    # Common Styles
    bold_font = Font(bold=True)
    header_font = Font(bold=True, size=12, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    # Column Widths
    ws.column_dimensions['A'].width = 30  # Employee
    ws.column_dimensions['B'].width = 25  # Item
    ws.column_dimensions['C'].width = 15  # Total Requests
    ws.column_dimensions['D'].width = 18  # Avg Days Between
    ws.column_dimensions['E'].width = 18  # Weekly Usage
    ws.column_dimensions['F'].width = 18  # Yearly Usage
    ws.column_dimensions['G'].width = 20  # Status

    # Company Header
    apply_centro_header(ws, "SUPPLY CONSUMPTION ANALYSIS", 7)

    # Table Headers (Row 7)
    headers = ["Employee", "Item", "Total Requests", "Avg Days Between", "Weekly Usage", "Yearly Usage", "Status"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border

    # Populating Data - starting from Row 8
    for row_idx, data in enumerate(data_rows, 8):
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            cell.alignment = left_align if col_idx <= 2 else center_align

    filename = "CONSUMPTION_REPORT.xlsx"
    wb.save(filename)
    return filename

def generate_purchase_request_excel(pr_id):
    """Generates a professional Excel Purchase Request form in Portrait orientation with thick borders."""
    from database import SessionLocal, PurchaseRequest, PurchaseItem
    from sqlalchemy.orm import joinedload
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from datetime import datetime
    import os

    with SessionLocal() as session:
        pr = session.query(PurchaseRequest).options(joinedload(PurchaseRequest.items)).get(pr_id)
        if not pr:
            raise ValueError("PR not found.")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PURCHASE REQUEST"

        # Define styles
        bold_font = Font(bold=True)
        title_font = Font(bold=True, size=18)
        company_font = Font(bold=True, size=14)
        pr_no_label_font = Font(bold=True, color="C00000") # Red
        
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        
        # THICKER BORDERS - using medium thickness
        thick_border = Border(
            left=Side(style='medium'), right=Side(style='medium'),
            top=Side(style='medium'), bottom=Side(style='medium')
        )
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        bottom_border = Border(bottom=Side(style='thin'))

        # Set Column Widths for Portrait - Balanced for 3 signature columns
        ws.column_dimensions['A'].width = 34  # Item Description
        ws.column_dimensions['B'].width = 12  # Purpose / Reason
        ws.column_dimensions['C'].width = 12  # For (Department/End-User)
        ws.column_dimensions['D'].width = 10  # Price
        ws.column_dimensions['E'].width = 10  # QTY
        ws.column_dimensions['F'].width = 10  # Unit
        ws.column_dimensions['G'].width = 14  # Total

        # Row heights
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 25
        ws.row_dimensions[9].height = 35  # Header row

        # Header Section
        apply_centro_header(ws, "PURCHASE REQUEST FORM", 7)

        # Metadata Section
        # Row 7 - Date and PR No. (Shifted from Row 3)
        ws['A7'] = "Date:"
        ws['A7'].font = bold_font
        ws.merge_cells('B7:C7')
        ws['B7'] = pr.request_date.strftime("%m/%d/%Y")
        ws['B7'].border = bottom_border
        
        ws['D7'] = "PR No.:"
        ws['D7'].font = pr_no_label_font
        ws['D7'].alignment = right_align
        ws.merge_cells('E7:G7')
        ws['E7'] = pr.pr_no
        ws['E7'].font = bold_font
        ws['E7'].border = bottom_border
        
        # Row 8 - Department
        ws['A8'] = "Department:"
        ws['A8'].font = bold_font
        ws.merge_cells('B8:G8')
        ws['B8'] = pr.department
        ws['B8'].border = bottom_border
        
        # Row 9 - End-User
        ws['A9'] = "End-User / Person Who Will Use:"
        ws['A9'].font = bold_font
        ws.merge_cells('B9:G9')
        ws['B9'] = pr.end_user or ""
        ws['B9'].border = bottom_border
        
        # Row 10 - Position
        ws['A10'] = "Position:"
        ws['A10'].font = bold_font
        ws.merge_cells('B10:G10')
        ws['B10'] = pr.position or ""
        ws['B10'].border = bottom_border

        # Row 11 - Source / Supplier
        ws['A11'] = "Source / Supplier:"
        ws['A11'].font = bold_font
        ws.merge_cells('B11:G11')
        ws['B11'] = pr.supplier or ""
        ws['B11'].border = bottom_border

        # Item Details Row
        ws.merge_cells('A12:G12')
        ws['A12'] = "ITEM DETAILS"
        ws['A12'].font = bold_font

        # Table Headers with THICK BORDERS - Row 13
        headers = ["Item Description", "Purpose / Reason", "For (Department / End-User)", "Price", "QTY", "Unit", "Total"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=13, column=col_num)
            cell.value = header
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = thick_border
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Table Content - starting from Row 14
        current_row = 14
        total_sum = 0.0
        for item in pr.items:
            for col_idx, val in enumerate([item.description, item.purpose, item.for_dept, item.price, item.qty, item.unit, item.total], 1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.border = thick_border
                
                if col_idx == 1: cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                elif col_idx == 2: cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                elif col_idx == 3: cell.alignment = Alignment(horizontal="left", vertical="center")
                elif col_idx == 4: 
                    cell.alignment = right_align
                    cell.number_format = '#,##0.00'
                elif col_idx == 5: cell.alignment = center_align
                elif col_idx == 6: cell.alignment = center_align
                elif col_idx == 7: 
                    cell.alignment = right_align
                    cell.number_format = '#,##0.00'
            total_sum += item.total
            current_row += 1

        # Total Line
        ws.merge_cells(f'A{current_row+1}:E{current_row+1}')
        ws[f'A{current_row+1}'] = "Estimated Total:"
        ws[f'A{current_row+1}'].font = bold_font
        ws[f'A{current_row+1}'].alignment = left_align
        
        ws[f'F{current_row+1}'] = "Php."
        ws[f'F{current_row+1}'].font = bold_font
        ws[f'F{current_row+1}'].alignment = right_align
        
        ws[f'G{current_row+1}'] = total_sum
        ws[f'G{current_row+1}'].font = bold_font
        ws[f'G{current_row+1}'].alignment = right_align
        ws[f'G{current_row+1}'].number_format = '#,##0.00'

        # Signature Section - Balanced into 3 equal widths (Col A, B-D, E-G)
        sig_row = current_row + 3
        # Headers
        ws[f'A{sig_row}'] = "Requested By:"
        ws[f'A{sig_row}'].font = bold_font
        ws[f'A{sig_row}'].alignment = center_align
        
        ws.merge_cells(f'B{sig_row}:D{sig_row}')
        ws[f'B{sig_row}'] = "Prepared By:"
        ws[f'B{sig_row}'].font = bold_font
        ws[f'B{sig_row}'].alignment = center_align
        
        ws.merge_cells(f'E{sig_row}:G{sig_row}')
        ws[f'E{sig_row}'] = "Approved By:"
        ws[f'E{sig_row}'].font = bold_font
        ws[f'E{sig_row}'].alignment = center_align
        
        # Names
        ws[f'A{sig_row+1}'] = "" # Space for signature/name
        ws[f'A{sig_row+1}'].alignment = center_align
        
        ws.merge_cells(f'B{sig_row+1}:D{sig_row+1}')
        ws[f'B{sig_row+1}'] = pr.prepared_by or ""
        ws[f'B{sig_row+1}'].alignment = center_align
        
        ws.merge_cells(f'E{sig_row+1}:G{sig_row+1}')
        ws[f'E{sig_row+1}'] = pr.approved_by or ""
        ws[f'E{sig_row+1}'].alignment = center_align
        
        # (Signature underlines removed)
        
        # (Labels row removed per user request)

        # Page Setup
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        
        ws.page_margins.left = 0.3
        ws.page_margins.right = 0.3
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        
        ws.print_options.horizontalCentered = True
        ws.print_title_rows = '1:9'
        
        ws.print_area = f"A1:G{sig_row+1}"

        filename = f"PURCHASE_REQUEST_{pr.pr_no}.xlsx"
        wb.save(filename)
        return filename

def get_frequency_category(avg_days):
    """Categorizes based on the average gap between requests."""
    if avg_days is None or avg_days >= 400: return "UNTIL DEFECTIVE"
    if avg_days < 2: return "Daily"
    if avg_days < 5: return "Twice a Week"
    if avg_days < 10: return "Weekly"
    if avg_days < 45: return "Monthly"
    if avg_days < 400: return "Annually"
    return "UNTIL DEFECTIVE"

def generate_high_frequency_report(data_rows):
    """Generates report for items with highest request frequency."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "High Frequency Items"
    
    # Apply standard Centro Header
    apply_centro_header(ws, "HIGH FREQUENCY ITEMS ANALYSIS", 4)

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Headers - Row 7
    headers = ["Item Name", "Total Quantity Requested", "Total Request Count", "Avg Frequency"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col)].width = 25

    # Data - starting from Row 8
    for row_idx, data in enumerate(data_rows, 8):
        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    filename = f"HIGH_FREQUENCY_ITEMS_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    return filename

def generate_pending_requests_report(data_rows):
    """Generates report for pending requests."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pending Requests"
    
    # Apply standard Centro Header
    apply_centro_header(ws, "OFFICIAL PENDING REQUESTS LOG", 6)

    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Headers - Row 7
    headers = ["Date", "Employee", "Item", "Quantity", "Area", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Data - starting from Row 8
    for row_idx, data in enumerate(data_rows, 8):
        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    filename = f"PENDING_REQUESTS_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    return filename

def generate_employee_behavior_report(data_rows):
    """Generates report for employee consumption patterns."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee behavior Analysis"
    
    # Apply standard Centro Header
    apply_centro_header(ws, "EMPLOYEE CONSUMPTION BEHAVIOR ANALYSIS", 6)

    header_fill = PatternFill(start_color="31869B", end_color="31869B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Headers - Row 7
    headers = ["Employee", "Item", "Total Qty", "Last Date", "Avg Days Between", "Frequency Classification"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col)].width = 25

    # Data - starting from Row 8
    for row_idx, data in enumerate(data_rows, 8):
        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    filename = f"EMPLOYEE_BEHAVIOR_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    return filename

def generate_inventory_report(report_type, data_rows):
    """Generates various types of inventory reports based on context."""
    from datetime import datetime
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Title Mapping
    titles = {
        "stock_summary": "STOCK SUMMARY REPORT",
        "low_stock": "LOW STOCK ITEMS (BELOW STANDARD)",
        "needs_restock": "URGENT RESTOCK NEEDS",
        "distribution": "INVENTORY DISTRIBUTION BY LOCATION",
        "custom_items": "CUSTOM ITEM SELECTION REPORT"
    }
    ws.title = report_type.replace("_", " ").title()
    current_title = titles.get(report_type, "Inventory Report")

    # Appply standard Centro Header
    apply_centro_header(ws, current_title, 5)

    # Header Styles
    header_fills = {
        "stock_summary": PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid"),
        "low_stock": PatternFill(start_color="E26B0A", end_color="E26B0A", fill_type="solid"),
        "needs_restock": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        "distribution": PatternFill(start_color="31869B", end_color="31869B", fill_type="solid")
    }
    header_fill = header_fills.get(report_type, header_fills["stock_summary"])
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Headers - Row 7
    if report_type == "distribution":
        headers = ["Item Name", "Location", "Quantity", "Unit", "Status"]
    else:
        headers = ["Item Name", "Description", "Actual Stock", "Standard Stock", "Unit"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col)].width = 25

    # Data - starting from Row 8
    for row_idx, data in enumerate(data_rows, 8):
        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    filename = f"INVENTORY_{report_type.upper()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    return filename

if __name__ == "__main__":
    generate_blank_form()
