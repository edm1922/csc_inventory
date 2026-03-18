import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os

FORM_FILENAME = "BLANK_REQUEST_FORM.xlsx"

def generate_blank_form():
    """Generates a perfectly formatted, printable blank Excel form for supply requests."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Supply Request Form"

    # Define common styles
    bold_font = Font(bold=True)
    header_font = Font(bold=True, size=14, color="FFFFFF")
    title_font = Font(bold=True, size=18)
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Column Widths
    ws.column_dimensions['A'].width = 5   # "#"
    ws.column_dimensions['B'].width = 35  # "Item Name"
    ws.column_dimensions['C'].width = 12  # "Quantity"
    ws.column_dimensions['D'].width = 15  # "Is Refill? (Y/N)"
    ws.column_dimensions['E'].width = 25  # "Frequency / Remarks"

    # Company Name Header
    ws.merge_cells('B1:E1')
    ws['B1'] = "CENTRO SERVICES COOPERATIVE"
    ws['B1'].font = Font(bold=True, size=24, color="1F497D") # Dark Blue
    ws['B1'].alignment = center_align

    # Form Title
    ws.merge_cells('B2:E2')
    ws['B2'] = "MATERIAL / SUPPLY REQUEST FORM"
    ws['B2'].font = title_font
    ws['B2'].alignment = center_align

    # Employee / Department Block (Top)
    metadata_labels = [
        ("B3", "Date:"), ("D3", "Shift:"),
        ("B4", "Employee Name:"), ("D4", "Employee Role:"),
        ("B5", "Department Area:"), ("D5", "Supervisor:")
    ]
    
    for cell_loc, label in metadata_labels:
        ws[cell_loc] = label
        ws[cell_loc].font = bold_font
        ws[cell_loc].alignment = left_align
        
        # Create a line to write on next to the label
        blank_cell = f"{chr(ord(cell_loc[0]) + 1)}{cell_loc[1]}"
        ws[blank_cell].border = Border(bottom=Side(style='thin'))

    # Table Headers (Row 8)
    headers = ["#", "Item Requested", "Quantity", "Is Refill? (Y/N)", "Frequency / Remarks"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border

    # Blank Rows for filling (Rows 9 to 25)
    for row_num in range(9, 26):
        # Auto-number the # column
        num_cell = ws.cell(row=row_num, column=1)
        num_cell.value = row_num - 8
        num_cell.alignment = center_align
        num_cell.border = thin_border
        
        # Add borders to blank cells
        for col_num in range(2, 6):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border

    # Signatures block at the bottom
    ws['B28'] = "Requested By:"
    ws['B28'].font = bold_font
    ws['C28'].border = Border(bottom=Side(style='thin'))
    ws.merge_cells('C28:E28')
    
    ws['B30'] = "Approved By:"
    ws['B30'].font = bold_font
    ws['C30'].border = Border(bottom=Side(style='thin'))
    ws.merge_cells('C30:E30')

    # Page Setup for Printing
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    
    # Set print margins
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75

    # Save the form
    wb.save(FORM_FILENAME)
    print(f"Printable form generated at: {os.path.abspath(FORM_FILENAME)}")

def generate_populated_report(employee_name, role, area, shift, supervisor, requests_data):
    """Generates a populated supply request history report for a specific employee."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Request History"

    # Define common styles
    bold_font = Font(bold=True)
    header_font = Font(bold=True, size=14, color="FFFFFF")
    title_font = Font(bold=True, size=18)
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Column Widths
    ws.column_dimensions['A'].width = 15  # "Date"
    ws.column_dimensions['B'].width = 30  # "Item Requested"
    ws.column_dimensions['C'].width = 12  # "Quantity"
    ws.column_dimensions['D'].width = 15  # "Is Refill?"
    ws.column_dimensions['E'].width = 25  # "Frequency"

    # Company Name Header
    ws.merge_cells('A1:E1')
    ws['A1'] = "CENTRO SERVICES COOPERATIVE"
    ws['A1'].font = Font(bold=True, size=24, color="1F497D")
    ws['A1'].alignment = center_align

    # Form Title
    ws.merge_cells('A2:E2')
    ws['A2'] = f"SUPPLY REQUEST HISTORY: {employee_name}"
    ws['A2'].font = title_font
    ws['A2'].alignment = center_align

    # Metadata Block
    ws['A4'] = "Employee Name:"; ws['B4'] = employee_name
    ws['A5'] = "Employee Role:"; ws['B5'] = role
    ws['A6'] = "Shift:";         ws['B6'] = shift
    ws['D4'] = "Area:";          ws['E4'] = area
    ws['D5'] = "Supervisor:";    ws['E5'] = supervisor or "N/A"

    for row in range(4, 7):
        for col in [1, 4]:
            ws.cell(row=row, column=col).font = bold_font

    # Table Headers (Row 8)
    headers = ["Date", "Item Requested", "Quantity", "Is Refill?", "Frequency / Remarks"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border

    # Populating Data
    for row_idx, data in enumerate(requests_data, 9):
        # data is expected to be (date_str, item_name, qty, refill_str, frequency)
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            cell.alignment = left_align if col_idx == 2 else center_align

    # Page Setup
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    
    filename = f"HISTORY_{employee_name.replace(' ', '_').replace(',', '')}.xlsx"
    wb.save(filename)
    return filename

def generate_consumption_report(data_rows):
    """Generates a summary Excel report for the Consumption & Usage Analysis."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Consumption Analysis"

    # Common Styles
    bold_font = Font(bold=True)
    header_font = Font(bold=True, size=12, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    # Column Widths
    ws.column_dimensions['A'].width = 30  # Employee
    ws.column_dimensions['B'].width = 25  # Item
    ws.column_dimensions['C'].width = 15  # Total Requests
    ws.column_dimensions['D'].width = 18  # Avg Days Between
    ws.column_dimensions['E'].width = 18  # Weekly Usage
    ws.column_dimensions['F'].width = 18  # Yearly Usage
    ws.column_dimensions['G'].width = 20  # Status

    # Company Header
    ws.merge_cells('A1:G1')
    ws['A1'] = "CENTRO SERVICES COOPERATIVE - SUPPLY CONSUMPTION ANALYSIS"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = center_align

    # Table Headers
    headers = ["Employee", "Item", "Total Requests", "Avg Days Between", "Weekly Usage", "Yearly Usage", "Status"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border

    # Populating Data
    for row_idx, data in enumerate(data_rows, 4):
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            cell.alignment = left_align if col_idx <= 2 else center_align

    filename = "CONSUMPTION_REPORT.xlsx"
    wb.save(filename)
    return filename

if __name__ == "__main__":
    generate_blank_form()
