import openpyxl
from openpyxl.styles import PatternFill

wb = openpyxl.load_workbook('sample.xlsx')
ws = wb.active

for row in ws.iter_rows(min_row=1, max_row=5):
    for cell in row:
        if cell.value:
            bg_color = cell.fill.fgColor.rgb if type(cell.fill) == PatternFill else 'None'
            font_color = cell.font.color.rgb if cell.font and cell.font.color else "None"
            font_prop = f"bold={cell.font.bold}, size={cell.font.size}, color={font_color}" if cell.font else "None"
            print(f"Cell {cell.coordinate}: value={cell.value}, bg={bg_color}, font={font_prop}")
